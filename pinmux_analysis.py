#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# SPDX-License-Identifier: GPL-2.0-only
#
# Pinmux Table Analysis for the iVot project
#
# Copyright (C) 2025 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#
import argparse
import json
import openpyxl
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from jsonschema import validate, ValidationError
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


##############################################################################
### Global Variable


CONFIG_SCHEMA = {
    '$schema': 'https://json-schema.org/draft/2020-12/schema',
    'type': 'object',
    'additionalProperties': False,
    'required': ['hide_row_parse', 'table_format', 'function', 'ignore', 'partition'],
    'properties': {
        'hide_row_parse': {'type': 'boolean'},
        'table_format': {
            'type': 'object',
            'additionalProperties': False,
            'required': ['active_ws', 'function', 'pad_name', 'ref_name'],
            'properties': {
                'active_ws': {'type': 'string'},
                'function': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'pattern'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'pattern': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        }
                    }
                },
                'pad_name': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'pattern'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'pattern': {'type': 'string'}
                    }
                },
                'ref_name': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'pattern'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'pattern': {'type': 'string'}
                    }
                }
            }
        },
        'function': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['subgroup', 'clock', 'custom'],
                    'properties': {
                        'subgroup': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'additionalProperties': False,
                                'required': ['pattern', 'name'],
                                'properties': {
                                    'pattern': {'type': 'string'},
                                    'name': {'type': 'string'}
                                }
                            }
                        },
                        'clock': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        },
                        'custom': {
                            'type': 'object',
                            'additionalProperties': False,
                            'patternProperties': {
                                r'\S+': {'type': 'string'}
                            }
                        }
                    }
                }
            }
        },
        'ignore': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {'type': 'string'}
            }
        },
        'partition': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {
                    'type': 'array',
                    'items': {'type': 'string'}
                }
            }
        }
    }
}

DIR_I_TAG = {'I', 'IO'}
DIR_O_TAG = {'O', 'IO'}
DIR_TAG = set(list(DIR_I_TAG) + list(DIR_O_TAG))


##############################################################################
### Data Structure


@dataclass
class Pin:
    func: str
    dir:  str
    pad:  str
    ref:  str


@dataclass
class SubGroup:
    pin_list: list[Pin] = field(default_factory=list)
    data_pin: list[Pin] = field(default_factory=list)
    clk_pin:  list[Pin] = field(default_factory=list)


@dataclass
class GroupData:
    repat:     list[re.Pattern] = field(default_factory=list)
    sub_name:  list[str]        = field(default_factory=list)
    clk_repat: list[re.Pattern] = field(default_factory=list)
    cus_pin:   dict[re.Pattern] = field(default_factory=dict)
    sub_group: dict[SubGroup]   = field(default_factory=dict)


##############################################################################
### Procedure


def parse_table(config: dict, workbook: Workbook, is_debug: bool=False) -> dict:
    """Parsing the pinmux table"""
    ws = workbook[config['table_format']['active_ws']]

    ### Get table format
    re_pat_list = []
    for pat in config['table_format']['function']['pattern']:
        re_pat_list.append(re.compile(pat))

    func_cidx_list = []
    for i, cell in enumerate(ws[config['table_format']['function']['rid']], start=1):
        for re_pat in re_pat_list:
            if re_pat.fullmatch(str(cell.value)):
                func_cidx_list.append(i)
                break
    func_cidx_list = [(x-1, x) for x in func_cidx_list]

    re_pat = re.compile(config['table_format']['pad_name']['pattern'])
    pad_cidx = None
    for i, cell in enumerate(ws[config['table_format']['pad_name']['rid']], start=1):
        value = str(cell.value).replace('\n', ' ').strip()
        if re_pat.fullmatch(value):
            pad_cidx = i
            break

    re_pat = re.compile(config['table_format']['ref_name']['pattern'])
    ref_cidx = None
    for i, cell in enumerate(ws[config['table_format']['ref_name']['rid']], start=1):
        value = str(cell.value).replace('\n', ' ').strip()
        if re_pat.fullmatch(value):
            ref_cidx = i
            break

    if is_debug:
        print('Function column index:', func_cidx_list)
        print('Pad name index:', pad_cidx)
        print('Ref name index:', ref_cidx)

    ### Get ignore dictionary
    ignore_dict = {}
    for gname, pat in config['ignore'].items():
        ignore_dict[gname] = {'active': False, 'repat': re.compile(pat)}

    if is_debug:
        print('\nIgnore dictionary: {')
        for gname, repat in ignore_dict.items():
            print('  {}: {},'.format(gname, repat['repat']))
        print('}')

    ### Get group format
    unknown_list = []
    group_dict = {}
    for gname, gconfig in config['function'].items():
        gdata = group_dict.setdefault(gname, GroupData())
        for sgroup in gconfig['subgroup']:
            gdata.repat.append(re.compile(sgroup['pattern']))
            gdata.sub_name.append(sgroup['name'])
        for pat in gconfig['clock']:
            gdata.clk_repat.append(re.compile(pat))
        for name, pat in gconfig['custom'].items():
            gdata.cus_pin[name] = re.compile(pat)

    if is_debug:
        debug_group_dict(group_dict, 'initial')

    ### Parsing table
    for ridx in range(1, ws.max_row+1):
        # row hidden check
        if ws.row_dimensions[ridx].hidden and not config['hide_row_parse']:
            print('check')
            continue

        for dir_cidx, func_cidx in func_cidx_list:
            # cell strike check
            if ws.cell(ridx, dir_cidx).font.strike == True:
                continue

            # ignore function check
            func_name, is_ignore = str(ws.cell(ridx, func_cidx).value), False
            for repat in ignore_dict.values():
                if repat['repat'].fullmatch(func_name):
                    repat['active'] = True
                    is_ignore = True
                    break
            if is_ignore:
                continue

            # parsing content
            direction = ws.cell(ridx, dir_cidx).value
            if direction is not None and str(direction).upper() in DIR_TAG:
                pin_data = Pin(
                    func=func_name,
                    dir=str(direction).upper(),
                    pad=ws.cell(ridx, pad_cidx).value,
                    ref=ws.cell(ridx, ref_cidx).value
                )
                
                if is_debug:
                    print(pin_data)

                is_unknown = True
                for gname, gdata in group_dict.items():
                    for pid, repat in enumerate(gdata.repat):
                        if (m := repat.fullmatch(pin_data.func)):
                            is_clk, is_unknown = False, False
                            for clk_repat in gdata.clk_repat:
                                if clk_repat.fullmatch(pin_data.func):
                                    is_clk = True
                                    break

                            sgname = repat.sub(gdata.sub_name[pid], pin_data.func)
                            sgdata = gdata.sub_group.setdefault(sgname, SubGroup())
                            if is_clk:
                                sgdata.clk_pin.append(pin_data)
                            else:
                                sgdata.data_pin.append(pin_data)
                            sgdata.pin_list.append(pin_data)

                if is_unknown:
                    unknown_list.append(pin_data)

    group_dict['unknown'] = unknown_list
    group_dict['ignore'] = ignore_dict
    if is_debug:
        debug_group_dict(group_dict, 'update')
    return group_dict


def print_group(group_dict: dict, out_fp):
    """Print the group result"""

    ### Sub function
    def _print_pin(pin_list: list, cpin_set: set|None=None, tabspace: int=0):
        func_len, pad_len, ref_len = 0, 0, 0
        for pin in pin_list:
            if (strlen := len(pin.func)) > func_len:
                func_len = strlen
            if (strlen := len(pin.pad)) > pad_len:
                pad_len = strlen
            if (strlen := len(pin.ref)) > ref_len:
                ref_len = strlen

        for pin in pin_list:
            if cpin_set is None:
                pin_type = ''
            elif pin.func in cpin_set:
                pin_type = '[C] '
            else:
                pin_type = '[D] '

            print('{}{}{} {:2} {} ({})'.format(
                    ' ' * tabspace,
                    pin_type,
                    pin.func.ljust(func_len), 
                    pin.dir, 
                    pin.pad.ljust(pad_len), 
                    pin.ref.ljust(ref_len)), 
                  file=out_fp)

    ### Print group result
    gname_len = 0
    for gname in group_dict.keys():
        if gname in {'unknown', 'ignore'}:
            continue
        if (new_len := len(gname)) > gname_len:
            gname_len = new_len
    gname_len += 6

    print('\n=== Category:\n', file=out_fp)
    for gname, gdata in group_dict.items():
        if gname in {'unknown', 'ignore'}:
            continue
        print(f'    {gname}: '.ljust(gname_len), end='', file=out_fp)
        msg = ''
        for sgname in gdata.sub_group.keys():
            msg += f'{sgname}, '
        print(msg[:-2], file=out_fp)

    if len(group_dict['unknown']):
        print('\n=== Unknown Function:\n', file=out_fp)
        # print the information of pins
        _print_pin(group_dict['unknown'], tabspace=4)

    active_ignore_dict, sgname_len = {}, 0
    for sgname, repat in group_dict['ignore'].items():
        if repat['active']:
            active_ignore_dict[sgname] = repat['repat']
            if (new_len := len(sgname)) > sgname_len:
                sgname_len = new_len

    if len(active_ignore_dict):
        print('\n=== Parsing Ignore:\n', file=out_fp)
        for sgname, repat in active_ignore_dict.items():
            print('    {}: {}'.format(sgname, repat.pattern), file=out_fp)

    print('\n=== Function:\n', file=out_fp)
    for gname, gdata in group_dict.items():
        if gname in {'unknown', 'ignore'}:
            continue

        for sgname, sgdata in gdata.sub_group.items():
            if len(sgdata.pin_list) == 0:
                continue
            print(f'- {sgname}:\n', file=out_fp)

            # print pin information
            cpin_set = set([pin.func for pin in sgdata.clk_pin])
            _print_pin(sgdata.pin_list, cpin_set=cpin_set, tabspace=4)
            print(file=out_fp)

            # print tool command
            cus_pin_dict = defaultdict(list)
            var_len = len(sgname) + 9

            ci_list, co_list = [], []
            for pin in sgdata.clk_pin:
                is_cus = False
                for cus_name, cus_repat in gdata.cus_pin.items():
                    if cus_repat.fullmatch(pin.func):
                        cus_pin_dict[cus_name].append(pin.pad)
                        is_cus = True
                        break

                if is_cus:
                    continue
                if pin.dir in DIR_I_TAG:
                    ci_list.append(pin.pad)
                if pin.dir in DIR_O_TAG:
                    co_list.append(pin.pad)

            if len(ci_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PADCLK_I').ljust(var_len),
                        ','.join(ci_list)), file=out_fp)

            if len(co_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PADCLK_O').ljust(var_len),
                        ','.join(co_list)), file=out_fp)

            di_list, do_list = [], []
            for pin in sgdata.data_pin:
                is_cus = False
                for cus_name, cus_repat in gdata.cus_pin.items():
                    if cus_repat.fullmatch(pin.func):
                        cus_pin_dict[cus_name].append(pin.pad)
                        is_cus = True
                        break

                if is_cus:
                    continue
                if pin.dir in DIR_I_TAG:
                    di_list.append(pin.pad)
                if pin.dir in DIR_O_TAG:
                    do_list.append(pin.pad)

            if len(di_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PORT_I').ljust(var_len),
                        ','.join(di_list)), file=out_fp)

            if len(do_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PORT_O').ljust(var_len),
                        ','.join(do_list)), file=out_fp)

            print(file=out_fp)

            if len(cus_pin_dict):
                var_len = 0
                for var_name in cus_pin_dict.keys():
                    if (new_len := len(var_name)) > var_len:
                        var_len = new_len
                var_len += 1

                for var_name, pin_list in cus_pin_dict.items():
                    print('{}set {} [get_ports {{{}}}]'.format(
                            ' ' * 4,
                            (sgname + '_' + var_name).ljust(var_len),
                            ','.join(pin_list)), file=out_fp)
                print(file=out_fp)


def debug_group_dict(group_dict: dict, status: str):
    print()
    print(f'Group({status}): {{')
    for gname, gdata in group_dict.items():
        if gname == 'unknown':
            print(f'  {gname}: [')
            for pin in gdata:
                print(f'    {pin},')
            print(f'  ],')
        elif gname == 'ignore':
            print(f'  {gname}: {{')
            for sgname, repat in gdata.items():
                print(f'    {sgname}: {repat},')
            print(f'  }},')
        else:
            print(f'  {gname}: {{')
            print(f'    repat:     {gdata.repat}')
            print(f'    sub_name:  {gdata.sub_name}')
            print(f'    clk_repat: {gdata.clk_repat}')
            print(f'    cus_pin:   {gdata.cus_pin}')
            for sgname, sgdata in gdata.sub_group.items():
                print(f'    {sgname}: {{')
                print(f'      data_pin: [')
                for pin in sgdata.data_pin:
                    print(f'        {pin},')
                print(f'      ],')
                print(f'      clk_pin: [')
                for pin in sgdata.clk_pin:
                    print(f'        {pin},')
                print(f'      ],')
                print(f'    }},')
            print(f'  }},')
    print('}\n')


##############################################################################
### Main


def create_argparse() -> argparse.ArgumentParser:
    """Create an argument parser."""
    parser = argparse.ArgumentParser(
                formatter_class=argparse.RawTextHelpFormatter,
                description='Pinmux Table Analysis for the iVot project')

    parser.add_argument('conf_fp', help="Configuration file") 
    parser.add_argument('table_fp', help="File path of the pinmux table") 
    parser.add_argument('-outfile', dest='out_fp', metavar='<file_path>', 
                            help="Set the output file path") 
    parser.add_argument('-debug', dest='is_debug', action='store_true', 
                                help="Debug mode")
    return parser


def main():
    """Main function."""
    parser = create_argparse()
    args = parser.parse_args()

    with open(args.conf_fp, 'r', encoding='utf-8') as json_fp:
        config = json.load(json_fp)
    try:
        validate(instance=config, schema=CONFIG_SCHEMA)
    except ValidationError as e:
        print(f'Error: JSON schema check fail.')
        print(e)
        exit(1)

    wb = openpyxl.load_workbook(args.table_fp)
    group_dict = parse_table(config, wb, is_debug=args.is_debug)

    if args.out_fp is None:
        print_group(group_dict, sys.stdout)
    else:
        with open(args.out_fp, 'w', encoding='utf-8') as fp:
            print_group(group_dict, fp)


if __name__ == '__main__':
    main()


