#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# SPDX-License-Identifier: GPL-2.0-only
#
# Pinmux Table Analyser for the iVot project
#
# Copyright (C) 2025 Yeh, Hsin-Hsien <yhh76227@gmail.com>
#
import argparse
import copy
import json
import math
import re
import sys
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from jsonschema import ValidationError, validate
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

##############################################################################
### Global Variable

CONFIG_SCHEMA = {
    '$schema': 'https://json-schema.org/draft/2020-12/schema',
    'type': 'object',
    'additionalProperties': False,
    'required': ['table_format', 'function', 'ignore', 'partition'],
    'properties': {
        'table_format': {
            'type': 'object',
            'additionalProperties': False,
            'required': ['active_tab', 'function', 'pad_name', 'ref_name', 'ignore'],
            'properties': {
                'active_tab': {'type': 'string'},
                'function': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'crange', 'pattern'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'crange': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        },
                        'pattern': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        }
                    }
                },
                'pad_name': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'pattern', 'style'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'pattern': {'type': 'string'},
                        'style': {
                            'type': 'string',
                            'enum': ['upper', 'lower', 'origin']
                        }
                    }
                },
                'ref_name': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['rid', 'pattern'],
                    'properties': {
                        'rid': {'type': 'integer'},
                        'pattern': {'type': 'string'}
                    }
                },
                'ignore': {
                    'type': 'object',
                    'additionalProperties': False,
                    'required': ['hide_row', 'font_strike'],
                    'patternProperties': {
                        'hide_row': {'type': 'boolean'},
                        'font_strike': {'type': 'boolean'},
                        'font_color_rgb': {'type': 'string'},
                        'font_color_index': {'type': 'integer'},
                        'font_color_theme': {'type': 'integer'},
                        r'_\S+': {}
                    }
                }
            }
        },
        'function': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {
                    'type': 'object',
                    'additionalProperties': False,
                    'patternProperties': {
                        'title': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        },
                        'sgroup': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'additionalProperties': False,
                                'required': ['pat', 'rep'],
                                'properties': {
                                    'pat': {'type': 'string'},
                                    'rep': {'type': 'string'},
                                    'ref': {
                                        'type': 'array',
                                        'items': {
                                            'type': 'object',
                                            'additionalProperties': False,
                                            'properties': {
                                                'p': {'type': 'string'},
                                                'r': {'type': 'string'}
                                            }
                                        }
                                    }
                                }
                            }
                        },
                        'sgorder': {
                            'type': 'object',
                            'additionalProperties': False,
                            'properties': {
                                'pat': {'type': 'string'},
                                'order': {
                                    'type': 'array',
                                    'items': {'type': 'integer'}
                                }
                            }
                        },
                        'clock': {
                            'type': 'array',
                            'items': {'type': 'string'}
                        },
                        'custom': {
                            'type': 'object',
                            'additionalProperties': False,
                            'patternProperties': {
                                r'\S+': {'type': 'string'}
                            }
                        },
                        r'_\S+': {}
                    }
                }
            }
        },
        'ignore': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {'type': 'string'}
            }
        },
        'partition': {
            'type': 'object',
            'additionalProperties': False,
            'patternProperties': {
                r'\S+': {
                    'type': 'array',
                    'items': {'type': 'string'}
                }
            }
        }
    }
}

DIR_I_TAG = {'I', 'IO'}
DIR_O_TAG = {'O', 'IO'}
DIR_TAG = set(list(DIR_I_TAG) + list(DIR_O_TAG))


##############################################################################
### Data Structure

@dataclass
class Pin:
    desc:   str
    dir:    str
    pad:    str
    ref:    str
    fname:  str
    index:  str
    font:   openpyxl.styles.fonts.Font|None = None


@dataclass
class SGName:
    pat:    re.Pattern
    rep:    str
    ref:    list[dict] = field(default_factory=list) 


@dataclass
class SGOrder:
    pat:    re.Pattern
    order:  list[int|str] = field(default_factory=list) 


@dataclass
class SGroup:
    plist:  list[Pin] = field(default_factory=list)
    dpin:   list[Pin] = field(default_factory=list)
    cpin:   list[Pin] = field(default_factory=list)


@dataclass
class GroupData:
    title:      list[re.Pattern]    = field(default_factory=list)
    sgname:     list[SGName]        = field(default_factory=list)
    sgorder:    SGOrder             = None
    clock:      list[re.Pattern]    = field(default_factory=list)
    custom:     dict[re.Pattern]    = field(default_factory=dict)
    sgroup:     defaultdict[SGroup] = field(init=False)
    def __post_init__(self):
        self.sgroup = defaultdict(SGroup)


@dataclass
class PartGroupDict:
    pat:    list[re.Pattern]       = field(default_factory=list)
    group:  defaultdict[GroupData] = field(init=False)
    def __post_init__(self):
        self.group = defaultdict(GroupData)


##############################################################################
### Procedure

def parse_table(config: dict, workbook: Workbook, is_debug: bool=False,
                is_dump_part: bool=False) -> dict:
    """Parsing the pinmux table"""
    ws = workbook[config['table_format']['active_tab']]

    ### Get table format
    repat_list = []
    for pat in config['table_format']['function']['pattern']:
        repat_list.append(re.compile(pat))

    len_crange = len(config['table_format']['function']['crange'])
    cidx_st, cidx_ed = 1, math.inf 
    if len_crange >= 1:
        cidx_st = column_index_from_string(config['table_format']['function']['crange'][0])
    if len_crange >= 2:
        cidx_ed = column_index_from_string(config['table_format']['function']['crange'][1])

    func_cidx_list = []
    for i, cell in enumerate(ws[config['table_format']['function']['rid']], start=1):
        if i < cidx_st:
            continue
        if i > cidx_ed:
            break
        for repat in repat_list:
            if repat.fullmatch(str(cell.value)):
                func_cidx_list.append(i)
                break
    func_cidx_list = [(x-1, x) for x in func_cidx_list]

    pad_repat = re.compile(config['table_format']['pad_name']['pattern'])
    pad_cidx = None
    for i, cell in enumerate(ws[config['table_format']['pad_name']['rid']], start=1):
        value = str(cell.value).replace('\n', ' ').strip()
        if pad_repat.fullmatch(value):
            pad_cidx = i
            break

    repat = re.compile(config['table_format']['ref_name']['pattern'])
    ref_cidx = None
    for i, cell in enumerate(ws[config['table_format']['ref_name']['rid']], start=1):
        value = str(cell.value).replace('\n', ' ').strip()
        if repat.fullmatch(value):
            ref_cidx = i
            break

    if is_debug:
        print('Function column index:', func_cidx_list)
        print('Pad name index:', pad_cidx)
        print('Ref name index:', ref_cidx)

    ### Get ignore dictionary
    ignore_dict = {}
    for gname, pat in config['ignore'].items():
        ignore_dict[gname] = {'active': False, 'repat': re.compile(pat)}

    if is_debug:
        print('\nIgnore dictionary: {')
        for gname, repat in ignore_dict.items():
            print('  {}: {},'.format(gname, repat['repat']))
        print('}')

    ### Get group format
    unknown_list = []
    group_dict = defaultdict(GroupData)
    for gname, gconfig in config['function'].items():
        if gname[0] == '_':
            continue
        gdata = group_dict[gname]
        if 'title' in gconfig:
            for pat in gconfig['title']:
                gdata.title.append(re.compile(pat))
        if 'sgroup' in gconfig:
            for sgroup in gconfig['sgroup']:
                sgname = SGName(pat=re.compile(sgroup['pat']), rep=sgroup['rep'])
                if 'ref' in sgroup:
                    for refine_pat in sgroup['ref']:
                        sgname.ref.append({'p': re.compile(refine_pat['p']), 
                                           'r': refine_pat['r']})
                gdata.sgname.append(sgname)
        if 'sgorder' in gconfig:
            gdata.sgorder = SGOrder(pat=re.compile(gconfig['sgorder']['pat']),
                                    order=copy.copy(gconfig['sgorder']['order']))
        if 'clock' in gconfig:
            for pat in gconfig['clock']:
                gdata.clock.append(re.compile(pat))
        if 'custom' in gconfig:
            for name, pat in gconfig['custom'].items():
                gdata.custom[name] = re.compile(pat)

    if is_debug:
        debug_group_dict(group_dict, 'initial')

    ### Get partition group format
    part_dict = defaultdict(PartGroupDict)
    for pname, pat_list in config['partition'].items():
        if pname[0] == '_':
            continue
        part_dict[pname].pat = [re.compile(x) for x in pat_list]
        part_dict[pname].group['unknown'] = []
        part_dict[pname].group['ignore'] = copy.deepcopy(ignore_dict)

    if is_debug:
        for pname, pdata in part_dict.items():
            print(f'{pname}: {pdata}')
        print()

    ### Parsing table
    func_name_dict = {}
    pad_all_set = set()
    pad_part_dict = {}
    config_ignore = config['table_format']['ignore']

    read_cell_str = lambda ridx, cidx: str(ws.cell(ridx, cidx).value).replace('\n', ' ').strip()

    for ridx in range(config['table_format']['function']['rid'], ws.max_row+1):
        # row hidden ignore check
        if ws.row_dimensions[ridx].hidden and not config_ignore['hide_row']:
            continue

        # get function name
        if pad_repat.fullmatch(read_cell_str(ridx, pad_cidx)):
            func_name_dict = {}
            for _, func_cidx in func_cidx_list:
                func_name_dict[func_cidx] = read_cell_str(ridx, func_cidx)
            continue

        for dir_cidx, func_cidx in func_cidx_list:
            # font strike ignore check
            if ws.cell(ridx, dir_cidx).font.strike and config_ignore['font_strike']:
                continue
            if ws.cell(ridx, func_cidx).font.strike and config_ignore['font_strike']:
                continue

            # font color ignore check
            fcolor = ws.cell(ridx, func_cidx).font.color
            if fcolor is not None:
                if ('font_color_rgb' in config_ignore and fcolor.type == 'rgb'
                    and fcolor.rgb.lower() == config_ignore['font_color_rgb'].lower()):
                    continue
                if ('font_color_index' in config_ignore and fcolor.type == 'indexed' 
                    and fcolor.indexed == config_ignore['font_color_index']):
                    continue
                if ('font_color_theme' in config_ignore and fcolor.type == 'theme' 
                    and fcolor.theme == config_ignore['font_color_theme']):
                    continue

            # ignore function check
            func_name = read_cell_str(ridx, func_cidx)
            ref_name = read_cell_str(ridx, ref_cidx)
            match config['table_format']['pad_name']['style']:
                case 'upper':
                    pad_name = read_cell_str(ridx, pad_cidx).upper()
                case 'lower':
                    pad_name = read_cell_str(ridx, pad_cidx).lower()
                case 'origin':
                    pad_name = read_cell_str(ridx, pad_cidx)

            is_ignore = False
            for gname, repat in ignore_dict.items():
                if repat['repat'].fullmatch(func_name):
                    repat['active'] = True
                    is_ignore = True
                    break
            if is_ignore:
                for pdata in part_dict.values():
                    for repat in pdata.pat:
                        if repat.fullmatch(pad_name):
                            pdata.group['ignore'][gname]['active'] = True
                continue

            # parsing content
            direction = ws.cell(ridx, dir_cidx).value
            if direction is not None and str(direction).upper() in DIR_TAG:
                pin_data = Pin(desc=func_name, 
                               dir=str(direction).upper(), 
                               pad=pad_name, 
                               ref=ref_name,
                               font=ws.cell(ridx, func_cidx).font,
                               fname=func_name_dict[func_cidx],
                               index=f'{get_column_letter(func_cidx)}{ridx}')

                pad_all_set.add(pad_name)
                
                if is_debug:
                    print(pin_data, '\n')

                is_unknown = True
                for gname, gdata in group_dict.items():
                    # active title check
                    if len(gdata.title) > 0:
                        title_hit = False
                        for repat in gdata.title:
                            if repat.fullmatch(func_name_dict[func_cidx]):
                                title_hit = True
                                break
                        if not title_hit:
                            continue

                    # is active sub-group pattern existed?
                    if len(gdata.sgname) == 0:
                        continue

                    for sgname_gen in gdata.sgname:
                        if (m := sgname_gen.pat.fullmatch(pin_data.desc)):
                            is_clk, is_unknown = False, False
                            for repat in gdata.clock:
                                if repat.fullmatch(pin_data.desc):
                                    is_clk = True
                                    break

                            # add to the group dictionary
                            sgname = sgname_gen.pat.sub(sgname_gen.rep, pin_data.desc)
                            for refine_pat in sgname_gen.ref:
                                sgname = refine_pat['p'].sub(refine_pat['r'], sgname)
                            sgdata = gdata.sgroup[sgname]
                            sgdata.plist.append(pin_data)
                            if is_clk:
                                sgdata.cpin.append(pin_data)
                            else:
                                sgdata.dpin.append(pin_data)

                            # check and add to the partition dictionary
                            is_check_done = False
                            for pname, pdata in part_dict.items():
                                for repat in pdata.pat:
                                    if repat.fullmatch(pin_data.pad):
                                        is_check_done = True
                                        pgdata = pdata.group[gname]
                                        if len(pgdata.sgroup) == 0:
                                            pgdata.sgorder = copy.deepcopy(gdata.sgorder)
                                            pgdata.custom = copy.deepcopy(gdata.custom)
                                        psgdata = pgdata.sgroup[sgname]
                                        psgdata.plist.append(pin_data)
                                        if is_clk:
                                            psgdata.cpin.append(pin_data)
                                        else:
                                            psgdata.dpin.append(pin_data)
                                        break

                                if is_check_done:
                                    is_check_done = False
                                    if pin_data.pad in pad_part_dict:
                                        pad_part_dict[pin_data.pad].add(pname)
                                    else:
                                        pad_part_dict[pin_data.pad] = set([pname])

                        if not is_unknown:
                            break
                    if not is_unknown:
                        break

                if is_unknown:
                    unknown_list.append(pin_data)
                    for pdata in part_dict.values():
                        for repat in pdata.pat:
                            if repat.fullmatch(pin_data.pad):
                                pdata.group['unknown'].append(pin_data)

    group_dict['unknown'] = unknown_list
    group_dict['ignore'] = ignore_dict

    if is_debug:
        debug_group_dict(group_dict, 'update')
        for pname, pdata in part_dict.items():
            debug_group_dict(pdata.group, f'update, {pname}')

    if is_dump_part and len(pad_part_dict) > 0:
        print()
        for pad_name, part_set in pad_part_dict.items():
            if len(part_set) > 1:
                print(f'Warning: PAD multi-assign: {pad_name} = {part_set}')
            pad_all_set.remove(pad_name)

        if len(pad_all_set) > 0:
            print('\nInformation: PAD non-assign:\n')

            strlen = 0
            for pad_name in pad_all_set:
                if (newlen := len(pad_name)) > strlen:
                    strlen = newlen
            strlen += 1

            msg = ''
            for i, pad_name in enumerate(pad_all_set):
                msg += f'{pad_name},'.rjust(strlen)
                if i % 10 == 9:
                    print(msg)
                    msg = ''
            if msg != '':
                print(msg)
            print()

    return group_dict, part_dict


def print_group(filename: str, sheetname: str, group_dict: dict, out_fp, scope: str|None=None):
    """Print the group result"""

    ### Sub function
    def _print_pin(pin_list: list, cpin_set: set|None=None, tabspace: int=0, is_known: bool=True):
        desc_len, pad_len, ref_len, func_len, idx_len = 5, 4, 7, 8, 5
        for pin in pin_list:
            desc_len = max(desc_len, len(pin.desc))
            pad_len = max(pad_len, len(pin.pad))
            ref_len = max(ref_len, len(pin.ref))
            func_len = max(func_len, len(pin.fname))

        # Type, Function, Dir.
        print('{}| {} | {} | {} | {} | {} |'.format(
            ' ' * tabspace,
            'Type'.ljust(5),
            'Desc.'.ljust(desc_len),
            'Dir.'.ljust(4),
            'Port'.ljust(pad_len),
            'PadCell'.ljust(ref_len)
        ), end='', file=out_fp)

        if is_known:
            print(' {} |'.format('Index'.ljust(idx_len)), file=out_fp)
        else:
            print(' {} | {} |'.format('Function'.ljust(func_len), 'Index'.ljust(idx_len)), file=out_fp)

        print('{}|-{}-|-{}-|-{}-|-{}-|-{}-|'.format(
            ' ' * tabspace,
            '-' * 5,
            '-' * desc_len,
            '-' * 4,
            '-' * pad_len,
            '-' * ref_len
        ), end='', file=out_fp)

        if is_known:
            print('-{}-|'.format('-' * idx_len), file=out_fp)
        else:
            print('-{}-|-{}-|'.format('-' * func_len, '-' * idx_len), file=out_fp)

        clk_val_list, dat_val_list = [], []

        for pin in pin_list:
            if cpin_set is None:
                pin_type = 'N/A'
            elif pin.desc in cpin_set:
                pin_type = 'Clock'
            else:
                pin_type = 'Data'

            val_list = clk_val_list if pin_type == 'Clock' else dat_val_list

            if is_known:
                val_list.append('{}| {} | {} | {} | {} | {} | {} |'.format(
                                    ' ' * tabspace,
                                    pin_type.ljust(5),
                                    pin.desc.ljust(desc_len),
                                    pin.dir.ljust(4),
                                    pin.pad.ljust(pad_len),
                                    pin.ref.ljust(ref_len),
                                    pin.index.ljust(idx_len)))
            else:
                val_list.append('{}| {} | {} | {} | {} | {} | {} | {} |'.format(
                                    ' ' * tabspace,
                                    pin_type.ljust(5),
                                    pin.desc.ljust(desc_len),
                                    pin.dir.ljust(4),
                                    pin.pad.ljust(pad_len),
                                    pin.ref.ljust(ref_len),
                                    pin.fname.ljust(func_len),
                                    pin.index.ljust(idx_len)))

        for val_str in clk_val_list:
            print(val_str, file=out_fp)
        for val_str in dat_val_list:
            print(val_str, file=out_fp)

    ### Print category
    table_dict, grp_len, func_len = {}, 5, 9

    for gname, gdata in group_dict.items():
        if gname in {'unknown', 'ignore'}:
            continue

        table_dict[gname] = []
        grp_len = max(grp_len, len(gname))
        func_len_tmp = 0

        # sub-group reorder
        if gdata.sgorder is not None:
            try:
                order_list = []
                for sgname in gdata.sgroup.keys():
                    toks = [sgname]
                    m = gdata.sgorder.pat.fullmatch(sgname)
                    for i in gdata.sgorder.order:
                        if m[i] in ('', None):
                            toks.append(0)
                        else:
                            toks.append(int(m[i]))
                    order_list.append(toks)

                gdata.sgorder.order = []
                for sgname, *_ in sorted(order_list, key=lambda x: x[1:]):
                    gdata.sgorder.order.append(sgname)
                    table_dict[gname].append(sgname)
                    func_len_tmp += len(sgname)
            except Exception as e:
                print(f'Error: incorrect sub-group reorder pattern, use the original order. ({gdata.sgorder})\n')
                traceback.print_exc()
                gdata.sgorder = None
                for sgname in gdata.sgroup.keys():
                    table_dict[gname].append(sgname)
                    func_len_tmp += len(sgname)
        else:
            for sgname in gdata.sgroup.keys():
                table_dict[gname].append(sgname)
                func_len_tmp += len(sgname)

        func_len = max(func_len, (func_len_tmp + len(table_dict[gname]) * 2))

    print('\n# File Information\n', file=out_fp)
    print('{}Filename:  {}'.format(' ' * 4, filename), file=out_fp)
    print('{}Sheetname: {}'.format(' ' * 4, sheetname), file=out_fp)
    if scope is None:
        print('{}Scope:     Top'.format(' ' * 4), file=out_fp)
    else:
        print('{}Scope:     Partition ({})'.format(' ' * 4, scope), file=out_fp)

    print('\n# Category\n', file=out_fp)
    print('{}| {} | {} |'.format(' ' * 4, 'Group'.ljust(grp_len), 'Functions'.ljust(func_len)), file=out_fp)
    print('{}|-{}-|-{}-|'.format(' ' * 4, ('-' * grp_len), ('-' * func_len)), file=out_fp)
    for gname, func_list in table_dict.items():
        func_str = ', '.join(func_list)
        print('{}| {} | {} |'.format(' ' * 4, gname.ljust(grp_len), func_str.ljust(func_len)), file=out_fp)

    ### Print unknown functions
    if len(group_dict['unknown']):
        print('\n# Unknown Function\n', file=out_fp)
        _print_pin(group_dict['unknown'], tabspace=4, is_known=False)

    ### Print ignore pattern
    table_dict, func_len, pat_len = {}, 8, 7

    for sgname, repat in group_dict['ignore'].items():
        if repat['active']:
            table_dict[sgname] = repat['repat']
            func_len = max(func_len, len(sgname))
            pat_len = max (pat_len, len(repat['repat'].pattern))

    if len(table_dict):
        print('\n# Parsing Ignore\n', file=out_fp)
        print('{}| {} | {} |'.format(' ' * 4, 'Function'.ljust(func_len), 'Pattern'.ljust(pat_len)), file=out_fp)
        print('{}|-{}-|-{}-|'.format(' ' * 4, ('-' * func_len), ('-' * pat_len)), file=out_fp)
        for fname, repat in table_dict.items():
            print('{}| {} | {} |'.format(' ' * 4, fname.ljust(func_len), repat.pattern.ljust(pat_len)), file=out_fp)

    ### Print function group
    print('\n# Function\n', file=out_fp)
    for gname, gdata in group_dict.items():
        if gname in {'unknown', 'ignore'}:
            continue

        # sub-group reorder
        if gdata.sgorder is not None:
            group_list = [(x, gdata.sgroup[x]) for x in gdata.sgorder.order]
        else:
            group_list = tuple(gdata.sgroup.items())

        for sgname, sgdata in group_list:
            if len(sgdata.plist) == 0:
                continue
            print(f'- {sgname}\n', file=out_fp)

            # print pin information
            cpin_set = set([pin.desc for pin in sgdata.cpin])
            _print_pin(sgdata.plist, cpin_set=cpin_set, tabspace=4)
            print(file=out_fp)

            # print tool command
            cus_pin_dict = defaultdict(list)
            var_len = len(sgname) + 9

            ci_list, co_list = [], []
            for pin in sgdata.cpin:
                is_cus = False
                for cus_name, cus_repat in gdata.custom.items():
                    if cus_repat.fullmatch(pin.desc):
                        cus_pin_dict[cus_name].append(pin.pad)
                        is_cus = True
                        break

                if is_cus:
                    continue
                if pin.dir in DIR_I_TAG:
                    ci_list.append(pin.pad)
                if pin.dir in DIR_O_TAG:
                    co_list.append(pin.pad)

            if len(ci_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PADCLK_I').ljust(var_len),
                        ' '.join(ci_list)), file=out_fp)

            if len(co_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PADCLK_O').ljust(var_len),
                        ' '.join(co_list)), file=out_fp)

            di_list, do_list = [], []
            for pin in sgdata.dpin:
                is_cus = False
                for cus_name, cus_repat in gdata.custom.items():
                    if cus_repat.fullmatch(pin.desc):
                        cus_pin_dict[cus_name].append(pin.pad)
                        is_cus = True
                        break

                if is_cus:
                    continue
                if pin.dir in DIR_I_TAG:
                    di_list.append(pin.pad)
                if pin.dir in DIR_O_TAG:
                    do_list.append(pin.pad)

            if len(di_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PORT_I').ljust(var_len),
                        ' '.join(di_list)), file=out_fp)

            if len(do_list):
                print('{}set {} [get_ports {{{}}}]'.format(
                        ' ' * 4,
                        (sgname + '_PORT_O').ljust(var_len),
                        ' '.join(do_list)), file=out_fp)

            print(file=out_fp)

            if len(cus_pin_dict):
                var_len = 0
                for var_name in cus_pin_dict.keys():
                    if (new_len := len(var_name)) > var_len:
                        var_len = new_len
                var_len += 1

                for var_name, pin_list in cus_pin_dict.items():
                    print('{}set {} [get_ports {{{}}}]'.format(
                            ' ' * 4,
                            (sgname + '_' + var_name).ljust(var_len),
                            ' '.join(pin_list)), file=out_fp)
                print(file=out_fp)


def debug_group_dict(group_dict: dict, status: str):
    print()
    print(f'Group({status}): {{')
    for gname, gdata in group_dict.items():
        if gname == 'unknown':
            print(f'  {gname}: [')
            for pin in gdata:
                print(f'    {pin},')
            print(f'  ],')
        elif gname == 'ignore':
            print(f'  {gname}: {{')
            for sgname, repat in gdata.items():
                print(f'    {sgname}: {repat},')
            print(f'  }},')
        else:
            print(f'  {gname}: {{')
            print(f'    title:     {gdata.title}')
            print(f'    sgname:    {gdata.sgname}')
            print(f'    sgorder:   {gdata.sgorder}')
            print(f'    clock:     {gdata.clock}')
            print(f'    custom:    {gdata.custom}')
            for sgname, sgdata in gdata.sgroup.items():
                print(f'    {sgname}: {{')
                print(f'      data_pin: [')
                for pin in sgdata.dpin:
                    print(f'        {pin},')
                print(f'      ],')
                print(f'      clk_pin: [')
                for pin in sgdata.cpin:
                    print(f'        {pin},')
                print(f'      ],')
                print(f'    }},')
            print(f'  }},')
    print('}\n')


##############################################################################
### Main

def create_argparse() -> argparse.ArgumentParser:
    """Create an argument parser."""
    parser = argparse.ArgumentParser(
                formatter_class=argparse.RawTextHelpFormatter,
                description='Pinmux Table Analyser for the iVot project')

    parser.add_argument('conf_fp', help='Configuration file') 
    parser.add_argument('table_fp', help='File path of the pinmux table') 
    parser.add_argument('-outfile', dest='out_fp', metavar='<file_path>', 
                            help='Set the output file path') 
    parser.add_argument('-dump_part', dest='is_dump_part', action='store_true', 
                            help='Dump partition result (effective if "-outfile" is defined)')
    parser.add_argument('-debug', dest='is_debug', action='store_true', 
                            help='Debug mode')
    return parser


def main():
    """Main function."""
    parser = create_argparse()
    args = parser.parse_args()

    with open(args.conf_fp, 'r', encoding='utf-8') as json_fp:
        config = json.load(json_fp)
    try:
        validate(instance=config, schema=CONFIG_SCHEMA)
    except ValidationError as e:
        print(f'Error: JSON schema check fail.')
        print(e)
        exit(1)

    is_dump_part = args.out_fp is not None and args.is_dump_part
    wb = openpyxl.load_workbook(args.table_fp)
    group_dict, part_dict = parse_table(config, wb, is_debug=args.is_debug, 
                                        is_dump_part=is_dump_part)

    filename = Path(args.table_fp).resolve()
    sheetname = config['table_format']['active_tab']

    if args.out_fp is None:
        print_group(filename, sheetname, group_dict, sys.stdout)
    else:
        with open(args.out_fp, 'w', encoding='utf-8') as fp:
            print_group(filename, sheetname, group_dict, fp)

    if args.is_dump_part:
        for pname, pdata in part_dict.items():
            if args.out_fp is None:
                print_group(filename, sheetname, group_dict, sys.stdout, pname)
            else:
                part_out_fp = Path(args.out_fp)
                part_out_fp = part_out_fp.parent / \
                                (part_out_fp.stem + f'_{pname}' + part_out_fp.suffix)
                with open(part_out_fp, 'w', encoding='utf-8') as fp:
                    print_group(filename, sheetname, pdata.group, fp, pname)


if __name__ == '__main__':
    main()


